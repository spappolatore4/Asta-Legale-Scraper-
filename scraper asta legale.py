import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import time
import pandas as pd
import os
import platform
import subprocess
from datetime import datetime
import ctypes
import requests
from bs4 import BeautifulSoup
import json
import re

try:
    import tkintermapview
    MAP_AVAILABLE = True
except ImportError:
    MAP_AVAILABLE = False


# ========== COMUNI PROVINCIA DI BRESCIA (lat, lon) ==========
COMUNI_BRESCIA = {
    # Città capoluogo
    "Brescia": (45.5417, 10.2118),
    # Cintura Brescia
    "Roncadelle": (45.5489, 10.1100), "Castegnato": (45.5622, 10.0600),
    "Ospitaletto": (45.5561, 10.0789), "Gussago": (45.5697, 10.1644),
    "Cellatica": (45.5672, 10.1833), "Collebeato": (45.5906, 10.2244),
    "Bovezzo": (45.5869, 10.2011), "Nave": (45.5906, 10.2883),
    "Concesio": (45.5978, 10.2217), "Botticino": (45.5289, 10.2467),
    "Rezzato": (45.5039, 10.3083), "Mazzano": (45.5294, 10.3194),
    "Nuvolento": (45.5467, 10.3694), "Nuvolera": (45.5494, 10.3522),
    "Prevalle": (45.5444, 10.3333), "Montirone": (45.5017, 10.1711),
    "Flero": (45.4861, 10.1872), "Poncarale": (45.4728, 10.1906),
    "Capriano del Colle": (45.4844, 10.2061), "Azzano Mella": (45.4656, 10.2022),
    "Borgosatollo": (45.4878, 10.2033), "San Zeno Naviglio": (45.4900, 10.2367),
    "Castenedolo": (45.4828, 10.2861), "Castel Mella": (45.4878, 10.1544),
    # Val Trompia
    "Villa Carcina": (45.6433, 10.1872), "Sarezzo": (45.6622, 10.1689),
    "Marcheno": (45.6794, 10.1694), "Gardone Val Trompia": (45.6867, 10.1844),
    "Lodrino": (45.7194, 10.1817), "Tavernole sul Mella": (45.7444, 10.1811),
    "Bovegno": (45.7761, 10.1861), "Pezzaze": (45.7656, 10.2361),
    "Irma": (45.7817, 10.2200), "Collio": (45.8356, 10.1972),
    "Brione": (45.6161, 10.2178),
    # Valle Sabbia
    "Vestone": (45.7294, 10.4178), "Idro": (45.7506, 10.4828),
    "Bagolino": (45.8156, 10.4689), "Lavenone": (45.7461, 10.4544),
    "Casto": (45.7194, 10.3872), "Agnosine": (45.7044, 10.4178),
    "Pertica Alta": (45.7783, 10.3683), "Pertica Bassa": (45.7644, 10.3544),
    "Roè Volciano": (45.6217, 10.4878), "Vobarno": (45.6383, 10.5044),
    "Gavardo": (45.5817, 10.4367), "Serle": (45.5856, 10.3689),
    # Lago di Garda
    "Desenzano del Garda": (45.4678, 10.5339), "Sirmione": (45.4867, 10.5961),
    "Lonato del Garda": (45.4656, 10.4811), "Pozzolengo": (45.4378, 10.5617),
    "Padenghe sul Garda": (45.5011, 10.5039), "Manerba del Garda": (45.5494, 10.5506),
    "Soiano del Lago": (45.5183, 10.5028), "Puegnago sul Garda": (45.5511, 10.5350),
    "Polpenazze del Garda": (45.5656, 10.5339), "Calvagese della Riviera": (45.5200, 10.4544),
    "Bedizzole": (45.5183, 10.4183), "Muscoline": (45.5511, 10.4700),
    "Salo'": (45.6078, 10.5183), "Gargnano": (45.6856, 10.6578),
    "Limone sul Garda": (45.8083, 10.7894), "Tremosine": (45.7739, 10.7206),
    "Tignale": (45.7272, 10.7017), "Toscolano-Maderno": (45.6367, 10.5989),
    "Moniga del Garda": (45.5417, 10.5183),
    # Franciacorta / Oglio
    "Iseo": (45.6583, 10.0528), "Corte Franca": (45.6217, 10.0289),
    "Erbusco": (45.5711, 10.0206), "Passirano": (45.5811, 10.0711),
    "Rovato": (45.5689, 10.0039), "Cologne": (45.5683, 9.9517),
    "Coccaglio": (45.5689, 9.9817), "Castrezzato": (45.5622, 9.9083),
    "Chiari": (45.5339, 9.9350), "Palazzolo sull'Oglio": (45.6022, 9.8817),
    "Paratico": (45.6528, 9.9544), "Sarnico": (45.6822, 9.9700),
    "Sulzano": (45.6844, 10.1028), "Sale Marasino": (45.7006, 10.1183),
    "Zone": (45.7339, 10.1183), "Marone": (45.7156, 10.1028),
    "Pisogne": (45.8039, 10.1028), "Tavernola Bergamasca": (45.7228, 10.0739),
    "Predore": (45.7044, 10.0150), "Castelli Calepio": (45.6711, 9.9361),
    "Urago d'Oglio": (45.5178, 9.8817), "Pontoglio": (45.5650, 9.8489),
    # Bassa Brescia
    "Leno": (45.3667, 10.2167), "Manerbio": (45.3589, 10.1367),
    "Ghedi": (45.4000, 10.2828), "Calvisano": (45.3500, 10.3500),
    "Remedello": (45.3661, 10.3856), "Carpenedolo": (45.3661, 10.4500),
    "Pralboino": (45.3194, 10.3200), "Isorella": (45.3844, 10.2833),
    "Gottolengo": (45.3844, 10.2689), "Bagnolo Mella": (45.4178, 10.1856),
    "Verolanuova": (45.3183, 10.0883), "Quinzano d'Oglio": (45.3178, 10.0183),
    "Orzinuovi": (45.4028, 9.9167), "Villachiara": (45.3483, 10.0333),
    "Visano": (45.3361, 10.3689), "Acquafredda": (45.3361, 10.3011),
    "Gambara": (45.2517, 10.2839), "Alfianello": (45.2633, 10.1783),
    "Pavone del Mella": (45.3106, 10.2006), "Milzano": (45.3106, 10.1744),
    "Robecco d'Oglio": (45.2522, 10.0639),
    # Valle Camonica
    "Darfo Boario Terme": (45.8850, 10.1867), "Artogne": (45.8656, 10.1683),
    "Pian Camuno": (45.8689, 10.1844), "Angolo Terme": (45.8861, 10.1694),
    "Breno": (45.9533, 10.3028), "Cedegolo": (45.9867, 10.3539),
    "Capo di Ponte": (46.0200, 10.3361), "Edolo": (46.1856, 10.3367),
    "Bienno": (45.9033, 10.2867), "Piancogno": (45.8994, 10.2361),
    "Borno": (45.9378, 10.2028), "Lozio": (45.9683, 10.2339),
    "Ossimo": (45.9183, 10.2533), "Berzo Inferiore": (45.9361, 10.2522),
    "Berzo Demo": (46.0367, 10.2700), "Corteno Golgi": (46.2028, 10.2039),
    "Incudine": (46.2189, 10.3194), "Temù": (46.2522, 10.3367),
    "Ponte di Legno": (46.2683, 10.5178), "Sonico": (46.1522, 10.3539),
    "Malonno": (46.1189, 10.3194), "Cevo": (46.0700, 10.2700),
    "Sellero": (46.0183, 10.2867), "Niardo": (45.9189, 10.3028),
    "Braone": (45.9222, 10.3194), "Cimbergo": (45.9867, 10.3033),
    "Ono San Pietro": (45.9700, 10.3194), "Costa Volpino": (45.8344, 10.0700),
    "Lovere": (45.8183, 10.0689), "Castro": (45.8161, 10.0878),
    "Rogno": (45.8189, 10.1200), "Gianico": (45.8844, 10.2033),
    "Prestine": (45.9367, 10.3061),
    # Alta Valle Sabbia / Giudicarie
    "Storo": (45.8522, 10.5706), "Bondone": (45.8011, 10.5133),
}


def _point_in_polygon(lat, lon, polygon):
    """Ray casting algorithm — verifica se (lat,lon) è dentro il poligono.
    polygon: lista di tuple (lat, lon).
    """
    n = len(polygon)
    inside = False
    j = n - 1
    for i in range(n):
        lat_i, lon_i = polygon[i]
        lat_j, lon_j = polygon[j]
        # Un raggio orizzontale (lat costante) da (lat,lon) verso destra:
        # l'edge (i,j) lo interseca se un vertice è sopra e l'altro sotto lat_test
        if ((lat_i > lat) != (lat_j > lat)) and \
           (lon < (lon_j - lon_i) * (lat - lat_i) / (lat_j - lat_i) + lon_i):
            inside = not inside
        j = i
    return inside


# ========== FINESTRA MAPPA INTERATTIVA ==========
class MappaZonaWindow:
    def __init__(self, parent, on_confirm, colors, existing_polygon=None):
        self.parent = parent
        self.on_confirm = on_confirm   # callback(comuni: list[str], polygon: list[tuple])
        self.C = colors
        self.polygon_points = []       # list of (lat, lon)
        self._markers = []
        self._polygon_obj = None
        self._lines = []
        self._existing_polygon = existing_polygon or []

        self.win = tk.Toplevel(parent)
        self.win.title("🗺  Selezione Zona — Provincia di Brescia")
        self.win.geometry("1150x780")
        self.win.minsize(900, 600)
        self.win.configure(bg=self.C["bg"])
        self.win.grab_set()
        self._build()
        # Ripristina zona precedente dopo che la mappa è pronta
        if self._existing_polygon and MAP_AVAILABLE:
            self.win.after(350, self._restore_existing_zone)

    # ── UI ──────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self.win, bg=self.C["accent"], height=62)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🗺  SELEZIONA LA ZONA DI INTERESSE — Provincia di Brescia",
                 font=("Georgia", 17, "bold"), fg=self.C["white"],
                 bg=self.C["accent"]).pack(expand=True, pady=14)

        # Istruzioni
        instr = tk.Frame(self.win, bg="#1A2535")
        instr.pack(fill=tk.X)
        tk.Label(instr,
                 text="  👆 Clicca sulla mappa per aggiungere punti  •  Almeno 3 punti per chiudere un'area  "
                      "•  La mappa mostra strade, vie, paesi e frazioni  •  Clicca ✅ CONFERMA per applicare il filtro",
                 font=("Segoe UI", 9), fg="#BDC3C7", bg="#1A2535",
                 anchor="w").pack(fill=tk.X, padx=12, pady=(7, 0))

        # ── Barra di ricerca ──────────────────────────────────────────────
        search_bar = tk.Frame(self.win, bg="#1A2535")
        search_bar.pack(fill=tk.X, padx=8, pady=(4, 2))

        tk.Label(search_bar, text="🔍 Cerca via / comune:",
                 font=("Segoe UI", 9, "bold"), fg=self.C["gold"],
                 bg="#1A2535").pack(side=tk.LEFT, padx=(4, 6))

        self.search_var = tk.StringVar()
        search_entry = tk.Entry(
            search_bar, textvariable=self.search_var,
            font=("Segoe UI", 10), bg=self.C["panel"],
            fg=self.C["text"], insertbackground=self.C["text"],
            relief="flat", width=38)
        search_entry.pack(side=tk.LEFT, padx=(0, 6), ipady=4)
        search_entry.bind("<Return>", lambda *_: self._do_search())

        tk.Button(search_bar, text="Cerca",
                  command=self._do_search,
                  bg="#2980B9", fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat",
                  padx=10, pady=4, cursor="hand2").pack(side=tk.LEFT)

        self.search_lbl = tk.Label(
            search_bar, text="← digita via, piazza, paese e premi Invio",
            font=("Segoe UI", 8, "italic"), fg=self.C["text_dim"],
            bg="#1A2535")
        self.search_lbl.pack(side=tk.LEFT, padx=10)

        # Body: mappa + pannello laterale
        body = tk.Frame(self.win, bg=self.C["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 0))

        # ── Pannello laterale destro ─────────────────────────────────────
        right = tk.Frame(body, bg=self.C["panel"], width=250)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(6, 0))
        right.pack_propagate(False)

        tk.Label(right, text="📍 Punti selezionati",
                 font=("Segoe UI", 10, "bold"), fg=self.C["gold"],
                 bg=self.C["panel"]).pack(anchor="w", padx=10, pady=(12, 4))

        self.pts_text = scrolledtext.ScrolledText(
            right, font=("Consolas", 8), bg="#0A0F18", fg="#A8D8A8",
            wrap=tk.WORD, state=tk.DISABLED, height=10)
        self.pts_text.pack(fill=tk.X, padx=8, pady=(0, 8))

        tk.Frame(right, bg=self.C["border"], height=1).pack(fill=tk.X, padx=8, pady=4)

        tk.Label(right, text="🏘 Comuni nella zona",
                 font=("Segoe UI", 10, "bold"), fg=self.C["gold"],
                 bg=self.C["panel"]).pack(anchor="w", padx=10, pady=(4, 4))

        self.comuni_text = scrolledtext.ScrolledText(
            right, font=("Consolas", 8), bg="#0A0F18", fg="#F0E68C",
            wrap=tk.WORD, state=tk.DISABLED)
        self.comuni_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        self.count_lbl = tk.Label(right, text="0 comuni selezionati",
                                  font=("Segoe UI", 9, "bold"), fg=self.C["success"],
                                  bg=self.C["panel"])
        self.count_lbl.pack(pady=(0, 8))

        # ── Mappa ────────────────────────────────────────────────────────
        map_frame = tk.Frame(body, bg=self.C["bg"])
        map_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        if not MAP_AVAILABLE:
            msg_frame = tk.Frame(map_frame, bg=self.C["bg"])
            msg_frame.pack(expand=True)
            tk.Label(msg_frame,
                     text="⚠️  Libreria mancante: tkintermapview\n\n"
                          "Installa con:\n\n  pip install tkintermapview\n\n"
                          "e riavvia l'applicazione.",
                     font=("Segoe UI", 13), fg=self.C["error"], bg=self.C["bg"],
                     justify="center").pack(expand=True)
        else:
            self.map_widget = tkintermapview.TkinterMapView(
                map_frame, corner_radius=0)
            self.map_widget.pack(fill=tk.BOTH, expand=True)
            # Tile server OpenStreetMap (già include strade, vie, paesi)
            self.map_widget.set_tile_server(
                "https://a.tile.openstreetmap.org/{z}/{x}/{y}.png", max_zoom=19)
            # Centro provincia di Brescia
            self.map_widget.set_position(45.62, 10.25)
            self.map_widget.set_zoom(10)
            self.map_widget.add_left_click_map_command(self._on_map_click)

        # ── Barra pulsanti ────────────────────────────────────────────────
        btn_bar = tk.Frame(self.win, bg=self.C["panel"])
        btn_bar.pack(fill=tk.X, padx=8, pady=6)

        bs = {"font": ("Segoe UI", 10, "bold"), "relief": "flat",
              "padx": 14, "pady": 8, "cursor": "hand2"}

        tk.Button(btn_bar, text="↩ Rimuovi Ultimo",
                  command=self._remove_last,
                  bg="#E67E22", fg="white", **bs).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_bar, text="🗑 Reset",
                  command=self._reset,
                  bg="#7F8C8D", fg="white", **bs).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_bar, text="🔍 Aggiorna Comuni",
                  command=self._refresh_comuni,
                  bg="#2980B9", fg="white", **bs).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_bar, text="🗺 Intera Provincia",
                  command=self._select_all_province,
                  bg="#8E44AD", fg="white", **bs).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_bar, text="✖ Annulla",
                  command=self.win.destroy,
                  bg="#555", fg="white", **bs).pack(side=tk.RIGHT, padx=4)
        tk.Button(btn_bar, text="✅ CONFERMA ZONA",
                  command=self._confirm,
                  bg=self.C["success"], fg="white",
                  font=("Segoe UI", 11, "bold"), relief="flat",
                  padx=18, pady=8, cursor="hand2").pack(side=tk.RIGHT, padx=4)

        # ── Status bar ────────────────────────────────────────────────────
        self.status_var = tk.StringVar(
            value="Clicca sulla mappa per aggiungere il primo punto della zona...")
        tk.Label(self.win, textvariable=self.status_var,
                 font=("Consolas", 8), fg=self.C["gold"], bg=self.C["bg"],
                 anchor="w").pack(fill=tk.X, padx=12, pady=(0, 4))

    # ── Logica mappa ────────────────────────────────────────────────────────
    def _on_map_click(self, coords):
        lat, lon = coords
        self.polygon_points.append((lat, lon))
        n = len(self.polygon_points)

        if MAP_AVAILABLE:
            m = self.map_widget.set_marker(lat, lon, text=f"P{n}",
                                           marker_color_circle=self.C["accent"],
                                           marker_color_outside=self.C["accent2"])
            self._markers.append(m)
            self._redraw_polygon()

        self._update_pts_list()
        self._refresh_comuni()
        self.status_var.set(
            f"Punto {n} aggiunto — ({lat:.4f}, {lon:.4f})"
            + ("   ← aggiungi altri punti" if n < 3 else f"   ✔ {n} punti — zona definita"))

    def _redraw_polygon(self):
        # Elimina vecchio poligono
        if self._polygon_obj:
            try:
                self._polygon_obj.delete()
            except Exception:
                pass
            self._polygon_obj = None

        pts = self.polygon_points
        if len(pts) >= 3 and MAP_AVAILABLE:
            closed = list(pts) + [pts[0]]
            try:
                self._polygon_obj = self.map_widget.set_polygon(
                    closed,
                    fill_color="red",
                    outline_color=self.C["accent"],
                    border_width=3,
                    name="zona")
            except Exception:
                pass  # versioni vecchie di tkintermapview potrebbero non supportare fill_color rgba

    def _remove_last(self):
        if not self.polygon_points:
            return
        self.polygon_points.pop()
        if self._markers and MAP_AVAILABLE:
            try:
                self._markers[-1].delete()
            except Exception:
                pass
            self._markers.pop()
        self._redraw_polygon()
        self._update_pts_list()
        self._refresh_comuni()
        self.status_var.set(f"Punto rimosso — {len(self.polygon_points)} punti rimanenti.")

    def _reset(self):
        self.polygon_points.clear()
        for m in self._markers:
            try:
                m.delete()
            except Exception:
                pass
        self._markers.clear()
        if self._polygon_obj:
            try:
                self._polygon_obj.delete()
            except Exception:
                pass
            self._polygon_obj = None
        self._update_pts_list()
        self._refresh_comuni()
        self.status_var.set("Zona resettata — Clicca sulla mappa per iniziare.")

    def _select_all_province(self):
        """Seleziona automaticamente la rettangolo che copre tutta la provincia."""
        self._reset()
        # Bounding box provincia di Brescia (leggermente allargata)
        corners = [
            (46.35, 9.82), (46.35, 10.80),
            (45.20, 10.80), (45.20, 9.82),
        ]
        for lat, lon in corners:
            self.polygon_points.append((lat, lon))
            if MAP_AVAILABLE:
                n = len(self.polygon_points)
                m = self.map_widget.set_marker(lat, lon, text=f"P{n}",
                                               marker_color_circle=self.C["accent"],
                                               marker_color_outside=self.C["accent2"])
                self._markers.append(m)
        self._redraw_polygon()
        self._update_pts_list()
        self._refresh_comuni()
        self.status_var.set("Intera provincia di Brescia selezionata.")

    # ── Ricerca indirizzo ───────────────────────────────────────────────────
    def _do_search(self):
        if not MAP_AVAILABLE:
            return
        query = self.search_var.get().strip()
        if not query:
            return
        self.search_lbl.config(text="⏳ Ricerca in corso...", fg=self.C["gold"])
        self.win.update_idletasks()
        # Geocodifica in thread separato per non bloccare la UI
        threading.Thread(target=self._geocode_and_move, args=(query,), daemon=True).start()

    def _geocode_and_move(self, query):
        """Chiama Nominatim e sposta la mappa (thread separato)."""
        for search_q in (f"{query}, provincia di Brescia, Italia", f"{query}, Italia"):
            try:
                resp = requests.get(
                    "https://nominatim.openstreetmap.org/search",
                    params={"q": search_q, "format": "json", "limit": 1},
                    headers={"User-Agent": "AstaLegaleScraper/2.0"},
                    timeout=8)
                data = resp.json()
                if data:
                    lat = float(data[0]["lat"])
                    lon = float(data[0]["lon"])
                    label = data[0].get("display_name", query).split(",")[0].strip()
                    self.win.after(0, lambda la=lat, lo=lon, lb=label:
                                   self._move_map_to(la, lo, lb))
                    return
            except Exception:
                pass
        self.win.after(0, lambda: self.search_lbl.config(
            text="✖ Indirizzo non trovato — riprova con nome diverso",
            fg=self.C["error"]))

    def _move_map_to(self, lat, lon, label):
        """Centra e fa zoom sulla posizione trovata (thread principale)."""
        try:
            self.map_widget.set_position(lat, lon)
            self.map_widget.set_zoom(15)
            self.search_lbl.config(
                text=f"✔ {label}  ({lat:.4f}, {lon:.4f})  — clicca sulla mappa per aggiungere punti",
                fg=self.C["success"])
        except Exception as e:
            self.search_lbl.config(text=f"✖ Errore spostamento: {e}", fg=self.C["error"])

    # ── Ripristino zona esistente ───────────────────────────────────────────
    def _restore_existing_zone(self):
        """Ricarica i punti e il poligono della sessione precedente."""
        if not self._existing_polygon:
            return
        self.polygon_points = list(self._existing_polygon)
        for i, (lat, lon) in enumerate(self.polygon_points, 1):
            try:
                m = self.map_widget.set_marker(
                    lat, lon, text=f"P{i}",
                    marker_color_circle=self.C["accent"],
                    marker_color_outside=self.C["accent2"])
                self._markers.append(m)
            except Exception:
                pass
        self._redraw_polygon()
        self._update_pts_list()
        self._refresh_comuni()
        # Centra la mappa sulla zona
        lats = [p[0] for p in self.polygon_points]
        lons = [p[1] for p in self.polygon_points]
        self.map_widget.set_position(sum(lats) / len(lats), sum(lons) / len(lons))
        self.status_var.set(
            f"Zona precedente ripristinata — {len(self.polygon_points)} punti  "
            "• Puoi modificarla o confermarla direttamente")

    # ── Pannello laterale ───────────────────────────────────────────────────
    def _update_pts_list(self):
        self.pts_text.config(state=tk.NORMAL)
        self.pts_text.delete(1.0, tk.END)
        for i, (lat, lon) in enumerate(self.polygon_points, 1):
            self.pts_text.insert(tk.END, f"P{i}: {lat:.4f}, {lon:.4f}\n")
        self.pts_text.config(state=tk.DISABLED)

    def _refresh_comuni(self):
        comuni = self._get_comuni_in_zona()
        self.comuni_text.config(state=tk.NORMAL)
        self.comuni_text.delete(1.0, tk.END)
        if comuni:
            for c in comuni:
                self.comuni_text.insert(tk.END, f"• {c}\n")
        else:
            self.comuni_text.insert(tk.END,
                                    "(Nessun comune rilevato\n nella zona corrente)")
        self.comuni_text.config(state=tk.DISABLED)
        self.count_lbl.config(text=f"{len(comuni)} comuni selezionati")

    def _get_comuni_in_zona(self):
        if len(self.polygon_points) < 3:
            return []
        return sorted(
            nome for nome, (lat, lon) in COMUNI_BRESCIA.items()
            if _point_in_polygon(lat, lon, self.polygon_points)
        )

    # ── Conferma ────────────────────────────────────────────────────────────
    def _confirm(self):
        if len(self.polygon_points) < 3:
            messagebox.showwarning(
                "Zona incompleta",
                "Aggiungi almeno 3 punti per definire una zona geografica.",
                parent=self.win)
            return
        comuni = self._get_comuni_in_zona()
        if not comuni:
            if not messagebox.askyesno(
                    "Nessun comune",
                    "Nessun comune rilevato nella zona selezionata.\n\n"
                    "I risultati potrebbero essere vuoti.\nConfermare comunque?",
                    parent=self.win):
                return
        self.on_confirm(comuni, list(self.polygon_points))
        self.win.destroy()

# ========== CONFIGURAZIONE URL ==========
BASE_URL = "https://www.astalegale.net/Immobili"
DEFAULT_PARAMS = {
    "categories": "residenziali",
    "luoghi": "bs",           # ← CORRETTO: bs = provincia Brescia
    "tribunale": "brescia"
}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Cache-Control': 'max-age=0',
    'Referer': 'https://www.astalegale.net/',
}


# ========== DEBUG ==========
def save_debug_html(html_text, page_num=1):
    debug_dir = "debug_html_aste"
    os.makedirs(debug_dir, exist_ok=True)
    fname = os.path.join(debug_dir, f"page_{page_num}.html")
    with open(fname, 'w', encoding='utf-8') as f:
        f.write(html_text)
    print(f"[DEBUG] HTML salvato: {fname}")
    return fname


def save_debug_structure(soup, page_num=1):
    debug_dir = "debug_html_aste"
    os.makedirs(debug_dir, exist_ok=True)
    fname = os.path.join(debug_dir, f"structure_{page_num}.txt")

    with open(fname, 'w', encoding='utf-8') as f:
        f.write(f"=== STRUTTURA PAGINA {page_num} ===\n\n")
        f.write(f"TITLE: {soup.title.string if soup.title else 'N/A'}\n\n")

        f.write("=== CONTEGGIO TAG ===\n")
        for tag in ['article', 'div', 'li', 'section', 'a', 'p', 'span', 'h1', 'h2', 'h3', 'h4']:
            f.write(f"<{tag}>: {len(soup.find_all(tag))}\n")

        f.write("\n=== CLASSI CSS PIU USATE (top 60) ===\n")
        class_counts = {}
        for el in soup.find_all(True):
            for c in el.get('class', []):
                class_counts[c] = class_counts.get(c, 0) + 1
        for cls, cnt in sorted(class_counts.items(), key=lambda x: -x[1])[:60]:
            f.write(f"  .{cls}: {cnt}x\n")

        f.write("\n=== RICERCA PATTERN CHIAVE ===\n")

        for pattern_name, pattern in [
            ("Tribunale", r'Tribunale'),
            ("Data asta", r'[Dd]ata.*[Aa]sta'),
            ("Prezzo/euro", r'€|\d{4,}'),
            ("Abitazione/Fabbricato", r'Abitazione|Fabbricato'),
            ("Via/indirizzo", r'^(Via|Piazza|Viale|Corso|Largo)'),
        ]:
            els = soup.find_all(string=re.compile(pattern))
            f.write(f"\n'{pattern_name}': {len(els)} elementi\n")
            for el in els[:8]:
                parent = el.parent
                grandparent = parent.parent if parent else None
                f.write(f"  <{parent.name} class='{parent.get('class', [])}'>")
                if grandparent:
                    f.write(f" dentro <{grandparent.name} class='{grandparent.get('class', [])}'>")
                f.write(f"\n  TESTO: '{el.strip()[:100]}'\n")

        f.write("\n=== PRIME 3 CARD CANDIDATE ===\n")
        # article
        articles = soup.find_all('article')
        if articles:
            f.write(f"\n[article] trovati: {len(articles)}\n")
            f.write("PRIMA CARD:\n")
            f.write(articles[0].prettify()[:4000])
        else:
            f.write("\nNessun <article> trovato\n")
            # cerca div con Tribunale
            for div in soup.find_all('div'):
                txt = div.get_text()
                if 'Tribunale' in txt and ('Data' in txt or 'Abitazione' in txt):
                    if 100 < len(txt) < 3000:
                        f.write(f"\nDiv candidato (class={div.get('class', [])}):\n")
                        f.write(div.prettify()[:4000])
                        break

    print(f"[DEBUG] Struttura salvata: {fname}")
    return fname


# ========== SCRAPER ==========
class AstaLegaleScraper:
    def __init__(self):
        self.headers = HEADERS
        self.stop_scraping = False
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        self.debug_mode = True
        self.total_results = None

    def build_url(self, page=1, categories="residenziali", luoghi="bs", tribunale="brescia"):
        params = f"?categories={categories}&luoghi={luoghi}&tribunale={tribunale}"
        if page > 1:
            params += f"&page={page}"
        return BASE_URL + params

    def get_page(self, url, page_num=1):
        for attempt in range(3):
            try:
                time.sleep(0.5 if attempt == 0 else 3)
                resp = self.session.get(url, timeout=30)
                print(f"  HTTP {resp.status_code} ({len(resp.text)} chars)")

                if resp.status_code == 200:
                    if self.debug_mode:
                        save_debug_html(resp.text, page_num)
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    if self.debug_mode and page_num == 1:
                        save_debug_structure(soup, page_num)
                    return soup, resp.text

                elif resp.status_code == 403:
                    print("  403 Forbidden")
                    return None, None
                elif resp.status_code == 429:
                    print("  Rate limited, attendo 15s...")
                    time.sleep(15)
            except Exception as e:
                print(f"  Tentativo {attempt+1} fallito: {e}")
                time.sleep(3)
        return None, None

    def count_total(self, soup):
        try:
            full_text = soup.get_text()
            m = re.search(r'(\d+)\s+RISULTATI', full_text, re.IGNORECASE)
            if m:
                return int(m.group(1))
        except:
            pass
        return None

    def has_next_page(self, soup, current_page):
        try:
            next_num = current_page + 1
            if soup.find('a', href=re.compile(rf'page={next_num}')):
                return True
            if soup.find('a', string=re.compile(r'next|success|›|»', re.I)):
                return True
            pagination = soup.find(class_=re.compile(r'pagination|pager', re.I))
            if pagination and str(next_num) in pagination.get_text():
                return True
        except:
            pass
        return False

    def find_cards(self, soup):
        # Strategia 0: Astalegale.net — div.card-body.accordion (struttura Vue SSR)
        cards = soup.find_all('div', class_=lambda c: c and 'card-body' in c and 'accordion' in c)
        if len(cards) >= 2:
            print(f"  Trovate {len(cards)} card via div.card-body.accordion (astalegale)")
            return cards

        # Strategia 0b: fallback con border-pvp (class specifica per PVP)
        cards = soup.find_all('div', class_=lambda c: c and 'border-pvp' in c and 'card-body' in c)
        if len(cards) >= 2:
            print(f"  Trovate {len(cards)} card via div.card-body.border-pvp")
            return cards

        # Strategia 1: <article>
        articles = soup.find_all('article')
        if len(articles) >= 2:
            print(f"  Trovate {len(articles)} card via <article>")
            return articles

        # Strategia 2: classi card-like (evita card-header/card-footer annidati)
        for p in ['listing', 'result', 'property', 'annuncio', 'auction', 'lot']:
            cards = soup.find_all(['div', 'li'], class_=re.compile(p, re.I))
            if len(cards) >= 2:
                print(f"  Trovate {len(cards)} card via class~={p}")
                return cards

        # Strategia 3: div che contengono sia "Tribunale" sia data/categoria
        # Filtra solo i div più piccoli che contengono tutto il necessario
        candidates = []
        seen_ids = set()
        for div in soup.find_all(['div', 'li', 'section']):
            txt = div.get_text()
            if ('Tribunale' in txt and 300 < len(txt) < 2500 and
                    re.search(r'Abitazione|Fabbricato|\d{2}/\d{2}/\d{4}', txt) and
                    re.search(r'Prezzo|€', txt)):
                div_id = id(div)
                if div_id not in seen_ids:
                    seen_ids.add(div_id)
                    candidates.append(div)

        if len(candidates) >= 2:
            print(f"  Trovate {len(candidates)} card via contenuto Tribunale")
            return candidates

        print("  Nessuna card trovata. Vedi debug_html_aste/structure_1.txt")
        return []

    def extract_from_card(self, card):
        data = {}
        try:
            card_text = card.get_text(separator='\n', strip=True)

            # CATEGORIA
            cat = "N/A"
            # Astalegale.net: div.fw-semibold.text-uppercase.text-pvp > span
            cat_el = card.find('div', class_=lambda c: c and 'text-uppercase' in c and 'text-pvp' in c)
            if cat_el:
                t = cat_el.get_text(strip=True)
                if any(kw in t.upper() for kw in ['ABITAZIONE', 'FABBRICATO', 'CIVILE', 'ECONOMICO', 'RURALE']):
                    cat = t.title()
            if cat == "N/A":
                for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'strong', 'b']:
                    el = card.find(tag)
                    if el:
                        t = el.get_text(strip=True)
                        if any(kw in t.upper() for kw in ['ABITAZIONE', 'FABBRICATO', 'CIVILE', 'ECONOMICO', 'RURALE']):
                            cat = t.title()
                            break
            if cat == "N/A":
                m = re.search(r'(Abitazione\s+di\s+tipo\s+\w+|Fabbricato[^\n]*)', card_text, re.I)
                if m:
                    cat = m.group(1).strip().title()
            data['Categoria'] = cat

            # INDIRIZZO
            indirizzo = "N/A"
            # Astalegale.net: div.card-text-title contiene solo l'indirizzo
            addr_el = card.find('div', class_=lambda c: c and 'card-text-title' in c)
            if addr_el:
                t = addr_el.get_text(strip=True)
                if t and 5 < len(t) < 200:
                    indirizzo = t
            if indirizzo == "N/A":
                # Fallback: cerca elementi il cui testo INIZIA con prefisso indirizzo
                for el in card.find_all(['p', 'span', 'div', 'address']):
                    t = el.get_text(strip=True)
                    if re.match(r'^(Via|Viale|Corso|Piazza|Largo|Str|Loc|Fraz|S\.|C\.da|Contrada|Cascina)', t, re.I):
                        if 5 < len(t) < 120:
                            indirizzo = t
                            break
            if indirizzo == "N/A":
                m = re.search(
                    r'((?:Cascina|Loc\.|Via|Viale|Corso|Piazza|Largo|Str\.?|Fraz\.?)\s+[^\n,]{2,60}(?:,\s*[^\n]{2,50})?)',
                    card_text, re.I)
                if m:
                    indirizzo = m.group(1).strip()
            data['Via / Indirizzo'] = indirizzo

            # CITTA
            citta = "N/A"
            # Astalegale.net: span.comune contiene il badge città con icona
            comune_el = card.find('span', class_='comune')
            if comune_el:
                t = comune_el.get_text(strip=True)
                # Rimuove eventuale testo dell'icona (es. " Lodrino" → "Lodrino")
                t_clean = re.sub(r'[^\w\s\'\-àèéìòùÀÈÉÌÒÙ]', '', t).strip()
                if t_clean:
                    citta = t_clean
            if citta == "N/A":
                # Fallback generico: cerca badge con icona location
                for el in card.find_all('span', class_=lambda c: c and 'rounded-pill' in c):
                    t = el.get_text(strip=True)
                    t_clean = re.sub(r'[^\w\s\'\-àèéìòùÀÈÉÌÒÙ]', '', t).strip()
                    if (2 < len(t_clean) < 35 and
                            not any(kw in t_clean.lower() for kw in
                                    ['tribunale', 'abitazione', 'fabbricato', 'prezzo',
                                     'offerta', 'risultati', 'immobili']) and
                            not re.search(r'\d{4,}', t_clean)):
                        citta = t_clean
                        break
            # Fallback: CAP + città nell'indirizzo
            if citta == "N/A" and indirizzo != "N/A":
                m = re.search(r'\d{5}\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)', indirizzo)
                if m:
                    citta = m.group(1)
            data['Città'] = citta

            # TRIBUNALE
            m = re.search(r'(Tribunale\s+(?:di\s+)?\w+)', card_text, re.I)
            data['Tribunale'] = m.group(1).strip() if m else "Tribunale di Brescia"

            # DATA ASTA (solo giorno/mese/anno, senza orario)
            data_asta = "N/A"
            m = re.search(
                r'[Dd]ata\s+[Aa]sta\s*:?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{4})',
                card_text)
            if m:
                data_asta = m.group(1).strip()
            else:
                m = re.search(r'(\d{2}/\d{2}/\d{4})', card_text)
                if m:
                    data_asta = m.group(1).strip()
            data['Data Asta'] = data_asta

            # LOTTO
            lotto = "N/A"
            m = re.search(r'[Ll]otto\s*[nN°.]?\s*:?\s*(\w+)', card_text)
            if m:
                lotto = m.group(1).strip()
            data['Lotto'] = lotto

            # PREZZI
            prezzo_base = "N/A"
            offerta_min = "N/A"

            m = re.search(
                r"(?:[Pp]rezzo\s+[Bb]ase|[Bb]ase\s+d['\u2019]?\s*[Aa]sta)\s*:?\s*\u20ac?\s*([\d\.]+(?:,\d{2})?)",
                card_text)
            if m:
                prezzo_base = self._fmt_price(m.group(1))

            m = re.search(
                r'[Oo]fferta\s+[Mm]inima\s*:?\s*€?\s*([\d\.]+(?:,\d{2})?)',
                card_text)
            if m:
                offerta_min = self._fmt_price(m.group(1))

            # Fallback generico: tutti i prezzi nella card
            if prezzo_base == "N/A":
                all_prices = re.findall(r'€\s*([\d\.]+(?:,\d{2})?)', card_text)
                if all_prices:
                    prezzo_base = self._fmt_price(all_prices[0])
                if len(all_prices) > 1:
                    offerta_min = self._fmt_price(all_prices[1])

            data['Prezzo Base'] = prezzo_base
            data['Offerta Minima'] = offerta_min

            # URL
            url = "N/A"
            link = card.find('a', href=True)
            if link:
                href = link['href']
                url = href if href.startswith('http') else f"https://www.astalegale.net{href}"
            data['URL'] = url

        except Exception as e:
            print(f"  Errore parsing card: {e}")

        return data

    def _fmt_price(self, s):
        if not s:
            return "N/A"
        clean = str(s).replace('.', '').replace(',', '.').strip()
        try:
            return f"€ {float(clean):,.0f}".replace(',', '.')
        except:
            return str(s)

    def scrape_detail(self, url):
        if not url or url == "N/A":
            return {}
        soup, _ = self.get_page(url, page_num=9999)
        if not soup:
            return {}
        detail = {}
        try:
            full_text = soup.get_text(separator='\n', strip=True)
            for label, key in [
                (r'[Pp]rezzo\s+[Bb]ase', 'Prezzo Base'),
                (r'[Oo]fferta\s+[Mm]inima', 'Offerta Minima'),
                (r'[Rr]ilancio\s+[Mm]inimo', 'Rilancio Minimo'),
            ]:
                m = re.search(rf'{label}\s*:?\s*€?\s*([\d\.]+(?:,\d{{2}})?)', full_text)
                if m:
                    detail[key] = self._fmt_price(m.group(1))

            m = re.search(r'[Ll]otto\s*[nN°.]?\s*:?\s*(\w+)', full_text)
            if m:
                detail['Lotto'] = m.group(1)

            m = re.search(r'[Dd]ata\s+[Aa]sta\s*:?\s*(\d{1,2}/\d{2}/\d{4}\s*[-–]\s*\d{2}:\d{2})', full_text)
            if m:
                detail['Data Asta'] = m.group(1)

            m = re.search(r'(?:Via|Viale|Corso|Piazza|Largo|Loc\.?\s|Fraz\.?\s)[^\n]{5,80}',
                          full_text, re.I)
            if m:
                detail['Via / Indirizzo'] = m.group(0).strip()
        except:
            pass
        return detail

    def scrape_all(self, categories="residenziali", luoghi="bs", tribunale="brescia",
                   progress_callback=None, deep_scrape=False):
        all_results = []
        page = 1
        max_pages = 100
        consecutive_empty = 0

        while page <= max_pages and not self.stop_scraping:
            url = self.build_url(page, categories, luoghi, tribunale)

            if progress_callback:
                progress_callback(f"Pagina {page}: {url}", page, len(all_results))

            soup, raw = self.get_page(url, page_num=page)

            if not soup:
                consecutive_empty += 1
                if progress_callback:
                    progress_callback(f"Pagina {page} non scaricata", page, len(all_results))
                if consecutive_empty >= 2:
                    break
                page += 1
                continue

            if page == 1:
                total = self.count_total(soup)
                self.total_results = total
                if total and progress_callback:
                    est_pages = ((total - 1) // 12) + 1
                    progress_callback(f"Totale: {total} annunci (~{est_pages} pagine)", page, len(all_results))

            cards = self.find_cards(soup)

            if not cards:
                consecutive_empty += 1
                if progress_callback:
                    progress_callback(f"Nessuna card a pagina {page}", page, len(all_results))
                if consecutive_empty >= 2:
                    break
                page += 1
                time.sleep(2)
                continue

            consecutive_empty = 0
            page_count = 0

            for card in cards:
                if self.stop_scraping:
                    break
                item = self.extract_from_card(card)
                if item:
                    if deep_scrape and item.get('URL') and item['URL'] != 'N/A':
                        detail = self.scrape_detail(item['URL'])
                        for k, v in detail.items():
                            if item.get(k) in (None, 'N/A', ''):
                                item[k] = v
                        time.sleep(1)
                    all_results.append(item)
                    page_count += 1

            if progress_callback:
                progress_callback(
                    f"Pagina {page}: {page_count} annunci (tot: {len(all_results)})",
                    page, len(all_results))

            if not self.has_next_page(soup, page):
                if progress_callback:
                    progress_callback("Ultima pagina raggiunta", page, len(all_results))
                break

            time.sleep(2)
            page += 1

        return all_results, page


# ========== BARRA PROGRESSO ==========
class AnimatedProgressBar(ttk.Progressbar):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self._anim_running = False

    def start_animation(self):
        self._anim_running = True
        self._animate()

    def stop_animation(self):
        self._anim_running = False

    def _animate(self):
        if self._anim_running:
            self['value'] = (self['value'] + 2) % (self['maximum'] + 1)
            self.after(60, self._animate)


# ========== APP ==========
class AstaLegaleApp:
    def __init__(self, root):
        self.root = root
        self.is_scraping = False
        self.start_time = None
        self.timer_running = False
        self.total_found = 0
        self.scraper = AstaLegaleScraper()
        self.deep_scrape_var = tk.BooleanVar(value=False)
        self.debug_var = tk.BooleanVar(value=True)
        # Zona geografica selezionata dalla mappa
        self.zona_comuni = []          # es. ["Brescia", "Gussago", ...]
        self.zona_polygon = []         # lista (lat, lon)

        self._setup_colors()
        self._setup_window()
        self._build_ui()
        self._show_welcome()

    def _setup_colors(self):
        self.C = {
            "bg": "#0F1923", "panel": "#1A2535", "card": "#1E2D42",
            "border": "#2A3F5F", "accent": "#C0392B", "accent2": "#E74C3C",
            "gold": "#F39C12", "success": "#27AE60", "warning": "#F1C40F",
            "error": "#E74C3C", "text": "#ECF0F1", "text_dim": "#7F8C8D",
            "white": "#FFFFFF",
        }

    def _setup_window(self):
        self.root.title("⚖️  AstaLegale.net Scraper — Aste Giudiziarie Brescia")
        self.root.geometry("1280x860")
        self.root.minsize(1100, 760)
        self.root.configure(bg=self.C["bg"])
        self.root.update_idletasks()
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"1280x860+{(sw-1280)//2}+{(sh-860)//2}")

    def _build_ui(self):
        self._make_header()
        content = tk.Frame(self.root, bg=self.C["bg"])
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)
        left = tk.Frame(content, bg=self.C["bg"])
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        right = tk.Frame(content, bg=self.C["bg"], width=430)
        right.pack(side=tk.RIGHT, fill=tk.Y)
        right.pack_propagate(False)
        self._make_search_section(left)
        self._make_controls(right)
        self._make_progress_section(right)
        self._make_stats_section(right)
        self._make_log_section(right)
        self._make_footer()

    def _card(self, parent):
        return tk.Frame(parent, bg=self.C["card"],
                        highlightbackground=self.C["border"], highlightthickness=1)

    def _section_title(self, parent, text):
        tk.Frame(parent, bg=self.C["accent"], height=3).pack(fill=tk.X)
        tk.Label(parent, text=text, font=("Segoe UI", 13, "bold"),
                 fg=self.C["accent2"], bg=self.C["card"]).pack(
            anchor="w", padx=18, pady=(12, 10))

    def _make_header(self):
        hdr = tk.Frame(self.root, bg=self.C["accent"], height=90)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        inner = tk.Frame(hdr, bg=self.C["accent"])
        inner.pack(expand=True)
        tk.Label(inner, text="⚖️  ASTALEGALE.NET SCRAPER",
                 font=("Georgia", 26, "bold"), fg=self.C["white"],
                 bg=self.C["accent"]).pack(pady=(14, 3))
        tk.Label(inner, text="Aste Giudiziarie Immobiliari — Provincia di Brescia",
                 font=("Georgia", 12, "italic"), fg="#FADBD8",
                 bg=self.C["accent"]).pack()

    def _make_search_section(self, parent):
        card = self._card(parent)
        card.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        self._section_title(card, "🔍  PARAMETRI DI RICERCA")

        form = tk.Frame(card, bg=self.C["card"])
        form.pack(fill=tk.X, padx=18, pady=(0, 15))

        fields = [
            ("Categoria:", "residenziali", "categories", "es: residenziali, commerciali"),
            ("Luogo (sigla prov.):", "bs", "luoghi", "bs = Brescia, mi = Milano"),
            ("Tribunale:", "brescia", "tribunale", "es: brescia, milano"),
        ]
        self.filter_vars = {}
        for i, (label, default, key, hint) in enumerate(fields):
            tk.Label(form, text=label, font=("Consolas", 10, "bold"),
                     fg=self.C["gold"], bg=self.C["card"]).grid(
                row=i*2, column=0, sticky="w", padx=(0, 10), pady=(8, 0))
            var = tk.StringVar(value=default)
            self.filter_vars[key] = var
            tk.Entry(form, textvariable=var, font=("Consolas", 11),
                     bg=self.C["panel"], fg=self.C["text"],
                     insertbackground=self.C["text"],
                     relief="flat", width=22).grid(
                row=i*2, column=1, sticky="w", pady=(8, 0))
            tk.Label(form, text=hint, font=("Segoe UI", 8),
                     fg=self.C["text_dim"], bg=self.C["card"]).grid(
                row=i*2+1, column=1, sticky="w", pady=(0, 2))

        tk.Frame(card, bg=self.C["border"], height=1).pack(fill=tk.X, padx=18, pady=8)

        # ── Filtro zona geografica (mappa interattiva) ─────────────────────
        zona_frame = tk.Frame(card, bg=self.C["card"])
        zona_frame.pack(fill=tk.X, padx=18, pady=(0, 4))

        tk.Label(zona_frame, text="📍 Zona geografica:",
                 font=("Consolas", 10, "bold"),
                 fg=self.C["gold"], bg=self.C["card"]).pack(side=tk.LEFT)

        tk.Button(zona_frame, text="🗺 Apri Mappa",
                  command=self._open_mappa,
                  bg=self.C["accent"], fg=self.C["white"],
                  font=("Segoe UI", 9, "bold"), relief="flat",
                  padx=10, pady=4, cursor="hand2").pack(side=tk.LEFT, padx=(10, 6))

        tk.Button(zona_frame, text="✖ Rimuovi Filtro",
                  command=self._clear_zona,
                  bg="#555", fg=self.C["white"],
                  font=("Segoe UI", 8), relief="flat",
                  padx=8, pady=4, cursor="hand2").pack(side=tk.LEFT)

        self.zona_lbl = tk.Label(
            card,
            text="Nessun filtro zona attivo — scraping su tutta la provincia",
            font=("Segoe UI", 8, "italic"),
            fg=self.C["text_dim"], bg=self.C["card"], anchor="w")
        self.zona_lbl.pack(fill=tk.X, padx=18, pady=(0, 4))

        tk.Frame(card, bg=self.C["border"], height=1).pack(fill=tk.X, padx=18, pady=8)

        opt = tk.Frame(card, bg=self.C["card"])
        opt.pack(fill=tk.X, padx=18, pady=(0, 6))
        tk.Checkbutton(opt,
                       text=" 🔬 Scraping approfondito (visita ogni annuncio — più lento ma dati completi)",
                       variable=self.deep_scrape_var,
                       font=("Segoe UI", 9), fg=self.C["text"], bg=self.C["card"],
                       activebackground=self.C["card"],
                       selectcolor=self.C["panel"]).pack(anchor="w")

        opt2 = tk.Frame(card, bg=self.C["card"])
        opt2.pack(fill=tk.X, padx=18, pady=(0, 12))
        tk.Checkbutton(opt2,
                       text=" 🐛 Salva HTML debug in debug_html_aste/ (utile se dati escono N/A)",
                       variable=self.debug_var,
                       font=("Segoe UI", 9), fg=self.C["text_dim"], bg=self.C["card"],
                       activebackground=self.C["card"],
                       selectcolor=self.C["panel"]).pack(anchor="w")

        tk.Frame(card, bg=self.C["border"], height=1).pack(fill=tk.X, padx=18, pady=8)

        info = tk.Frame(card, bg=self.C["card"])
        info.pack(fill=tk.X, padx=18, pady=(0, 15))
        tk.Label(info, text="📋  CAMPI ESTRATTI:",
                 font=("Segoe UI", 10, "bold"), fg=self.C["gold"],
                 bg=self.C["card"]).pack(anchor="w")

        campi = ["Via / Indirizzo", "Città", "Categoria",
                 "Tribunale", "Lotto", "Prezzo Base",
                 "Offerta Minima", "Data Asta", "URL"]
        row_f = tk.Frame(info, bg=self.C["card"])
        row_f.pack(fill=tk.X, pady=(5, 0))
        for i, nome in enumerate(campi):
            tk.Label(row_f, text=f"● {nome}",
                     font=("Segoe UI", 9), fg=self.C["text_dim"],
                     bg=self.C["card"]).grid(
                row=i//3, column=i%3, sticky="w", padx=(0, 20), pady=2)

    def _make_controls(self, parent):
        card = self._card(parent)
        card.pack(fill=tk.X, pady=(0, 10))
        self._section_title(card, "🚀  AVVIA SCRAPING")

        tk.Label(card,
                 text="Aste giudiziarie residenziali\nprovincia di Brescia",
                 font=("Segoe UI", 9), fg=self.C["text_dim"],
                 bg=self.C["card"], justify="center").pack(pady=(0, 10))

        self.start_btn = tk.Button(
            card, text="⚖️  AVVIA SCRAPING",
            command=self._start,
            bg=self.C["accent"], fg=self.C["white"],
            font=("Segoe UI", 12, "bold"),
            relief="flat", padx=30, pady=12,
            cursor="hand2",
            activebackground=self.C["accent2"],
            activeforeground="white")
        self.start_btn.pack(pady=(0, 8))

        self.stop_btn = tk.Button(
            card, text="⏹  FERMA",
            command=self._stop,
            bg="#555", fg=self.C["white"],
            font=("Segoe UI", 10, "bold"),
            relief="flat", padx=20, pady=8,
            state=tk.DISABLED,
            activebackground="#333",
            activeforeground="white")
        self.stop_btn.pack(pady=(0, 8))

        tk.Button(card, text="🐛 Apri Debug HTML",
                  command=self._open_debug,
                  bg=self.C["panel"], fg=self.C["text_dim"],
                  font=("Segoe UI", 8), relief="flat",
                  padx=10, pady=4).pack(pady=(0, 14))

    def _make_progress_section(self, parent):
        card = self._card(parent)
        card.pack(fill=tk.X, pady=(0, 10))
        self._section_title(card, "📊  PROGRESSO")
        self.status_var = tk.StringVar(value="In attesa...")
        tk.Label(card, textvariable=self.status_var,
                 font=("Segoe UI", 9), fg=self.C["text"],
                 bg=self.C["card"], wraplength=390).pack(padx=14, pady=(0, 8))
        self.pbar = AnimatedProgressBar(card, maximum=100, length=390, mode='indeterminate')
        self.pbar.pack(padx=14, pady=(0, 8))
        self.quick_stats = tk.Label(card, text="Pagine: 0 | Annunci trovati: 0",
                                    font=("Consolas", 9), fg=self.C["gold"],
                                    bg=self.C["card"])
        self.quick_stats.pack(pady=(0, 12))

    def _make_stats_section(self, parent):
        card = self._card(parent)
        card.pack(fill=tk.X, pady=(0, 10))
        self._section_title(card, "⏱  STATISTICHE")
        self.time_lbl = tk.Label(card, text="🕐 Tempo: 00:00:00",
                                 font=("Consolas", 10), fg=self.C["text"],
                                 bg=self.C["card"])
        self.time_lbl.pack(pady=(0, 4))
        self.speed_lbl = tk.Label(card, text="⚡ Velocità: 0 annunci/min",
                                  font=("Consolas", 10), fg=self.C["text"],
                                  bg=self.C["card"])
        self.speed_lbl.pack(pady=(0, 12))

    def _make_log_section(self, parent):
        card = self._card(parent)
        card.pack(fill=tk.BOTH, expand=True)
        hdr = tk.Frame(card, bg=self.C["card"])
        hdr.pack(fill=tk.X, padx=14, pady=(10, 4))
        tk.Label(hdr, text="📝 LOG ATTIVITÀ",
                 font=("Segoe UI", 11, "bold"), fg=self.C["accent2"],
                 bg=self.C["card"]).pack(side=tk.LEFT)
        tk.Button(hdr, text="🗑 Pulisci", command=self._clear_log,
                  bg=self.C["border"], fg=self.C["text"],
                  font=("Segoe UI", 8), relief="flat",
                  padx=8, pady=2).pack(side=tk.RIGHT)
        self.log_text = scrolledtext.ScrolledText(
            card, font=("Consolas", 8),
            bg="#0A0F18", fg="#A8D8A8",
            insertbackground="white",
            wrap=tk.WORD, state=tk.DISABLED, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 14))

    def _make_footer(self):
        footer = tk.Frame(self.root, bg=self.C["panel"], height=40)
        footer.pack(fill=tk.X)
        footer.pack_propagate(False)
        tk.Button(footer, text="📁 Apri Risultati",
                  command=self._open_folder,
                  bg=self.C["panel"], fg=self.C["text_dim"],
                  font=("Segoe UI", 9), relief="flat",
                  padx=14, pady=5).pack(side=tk.LEFT, padx=14, pady=6)
        tk.Button(footer, text="🐛 Apri Debug",
                  command=self._open_debug,
                  bg=self.C["panel"], fg=self.C["text_dim"],
                  font=("Segoe UI", 9), relief="flat",
                  padx=14, pady=5).pack(side=tk.LEFT, padx=0, pady=6)
        tk.Label(footer, text="AstaLegale Scraper v2.0 | Python + Tkinter",
                 font=("Segoe UI", 8), fg=self.C["text_dim"],
                 bg=self.C["panel"]).pack(side=tk.RIGHT, padx=14, pady=12)

    def _show_welcome(self):
        self._log("""
⚖️  ASTALEGALE SCRAPER v2.0

✅ URL CORRETTO: luoghi=bs (sigla provincia Brescia)

🔧 SE I DATI ESCONO N/A:
   Il scraper salva automaticamente l'HTML grezzo in:
   📁 debug_html_aste/page_1.html     ← HTML completo
   📁 debug_html_aste/structure_1.txt ← analisi classi CSS

   Mandami il file structure_1.txt e aggiusto i selettori!

🚀 Premi AVVIA SCRAPING per iniziare.
""", "INFO")

    def _start(self):
        if self.is_scraping:
            messagebox.showwarning("Attenzione", "Scraping già in corso!")
            return
        if not messagebox.askyesno("Conferma",
                                   "Avviare scraping aste giudiziarie?\n\n"
                                   "Verranno salvati file di debug\n"
                                   "in debug_html_aste/ per verifica."):
            return

        self.is_scraping = True
        self.start_time = time.time()
        self.timer_running = True
        self.total_found = 0
        self.scraper.stop_scraping = False
        self.scraper.debug_mode = self.debug_var.get()

        self.start_btn.config(state=tk.DISABLED, text="⏳ Scraping...")
        self.stop_btn.config(state=tk.NORMAL, bg=self.C["error"])
        self.pbar.config(mode='indeterminate')
        self.pbar.start(30)

        self._update_timer()
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            cats = self.filter_vars['categories'].get()
            luoghi = self.filter_vars['luoghi'].get()
            tribunale = self.filter_vars['tribunale'].get()
            deep = self.deep_scrape_var.get()

            self._log(f"🚀 Avvio scraping", "SUCCESS")
            self._log(f"   categories={cats}, luoghi={luoghi}, tribunale={tribunale}", "INFO")
            if deep:
                self._log("   🔬 Modalità approfondita ATTIVA", "INFO")
            if self.zona_comuni:
                self._log(
                    f"   🗺 Filtro zona attivo: {len(self.zona_comuni)} comuni "
                    f"({', '.join(self.zona_comuni[:4])}"
                    f"{' ...' if len(self.zona_comuni) > 4 else ''})", "INFO")

            def cb(msg, page, found):
                self.total_found = found
                self._log(f"  {msg}", "INFO")
                self.root.after(0, lambda: self.quick_stats.config(
                    text=f"Pagine: {page} | Annunci trovati: {found}"))
                self.root.after(0, lambda: self.status_var.set(msg[:65]))

            results, pages = self.scraper.scrape_all(
                cats, luoghi, tribunale,
                progress_callback=cb,
                deep_scrape=deep)

            # ── Filtro zona geografica ──────────────────────────────────────
            if results and self.zona_comuni:
                prima = len(results)
                comuni_lower = {c.lower() for c in self.zona_comuni}
                results = [
                    r for r in results
                    if str(r.get('Città', '')).strip().lower() in comuni_lower
                ]
                self._log(
                    f"🗺 Filtro zona: {prima} → {len(results)} annunci "
                    f"(rimossi {prima - len(results)} fuori zona)", "INFO")

            self.total_found = len(results)

            if self.is_scraping:
                if results:
                    self._log(f"✅ Completato! {self.total_found} aste in {pages} pagine", "SUCCESS")
                    self.root.after(0, lambda: self.status_var.set(
                        f"✅ {self.total_found} aste trovate"))
                    self._save(results)
                else:
                    self._log("⚠️ Nessun risultato. Vedi debug_html_aste/structure_1.txt", "WARNING")
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Nessun risultato",
                        "Nessun annuncio estratto.\n\n"
                        "File debug salvati in: debug_html_aste/\n\n"
                        "Mandami il file structure_1.txt per\n"
                        "aggiustare i selettori!"))
        except Exception as e:
            import traceback
            self._log(f"❌ ERRORE: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self.root.after(0, lambda: messagebox.showerror("Errore", str(e)))
        finally:
            self.is_scraping = False
            self.timer_running = False
            self.scraper.stop_scraping = True
            self.root.after(0, self._reset_ui)

    def _stop(self):
        if not self.is_scraping:
            return
        if messagebox.askyesno("Ferma", "Vuoi interrompere lo scraping?"):
            self.is_scraping = False
            self.timer_running = False
            self.scraper.stop_scraping = True
            self.status_var.set("⏹ Interrotto")
            self._log("🛑 SCRAPING INTERROTTO", "WARNING")
            self._reset_ui()

    def _reset_ui(self):
        self.pbar.stop()
        self.pbar.stop_animation()
        self.pbar.config(mode='determinate')
        self.pbar['value'] = 100 if self.total_found > 0 else 0
        self.start_btn.config(state=tk.NORMAL, text="⚖️  AVVIA SCRAPING")
        self.stop_btn.config(state=tk.DISABLED, bg="#555")

    def _save(self, results):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            out_dir = "risultati_aste"
            os.makedirs(out_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fpath = os.path.join(out_dir, f"aste_brescia_{ts}.xlsx")

            df = pd.DataFrame(results)
            col_order = ['Categoria', 'Via / Indirizzo', 'Città',
                         'Tribunale', 'Lotto', 'Prezzo Base',
                         'Offerta Minima', 'Data Asta', 'URL']
            if 'Rilancio Minimo' in df.columns:
                col_order.insert(7, 'Rilancio Minimo')
            for c in col_order:
                if c not in df.columns:
                    df[c] = "N/A"
            df = df.reindex(columns=col_order).fillna("N/A")

            RED = "C0392B"
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Aste Brescia', index=False)
                ws = writer.sheets['Aste Brescia']

                for col, w in zip('ABCDEFGHI', [30, 38, 18, 22, 10, 18, 18, 22, 10]):
                    ws.column_dimensions[col].width = w

                # Header
                hfill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = hfill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin

                # Dati
                alt_fill = PatternFill(start_color="FDF2F8", end_color="FDF2F8", fill_type="solid")
                center = Alignment(horizontal="center", vertical="center")
                left_a = Alignment(horizontal="left", vertical="center")
                gold_f = Font(bold=True, color="E67E22", size=10)
                green_f = Font(bold=True, color="27AE60", size=10)

                max_row = ws.max_row
                for r in range(2, max_row + 1):
                    ws.row_dimensions[r].height = 22
                    for c_idx, col_name in enumerate(col_order, 1):
                        cell = ws.cell(row=r, column=c_idx)
                        cell.border = thin
                        if col_name == 'Prezzo Base':
                            cell.font = gold_f
                            cell.alignment = center
                        elif col_name == 'Offerta Minima':
                            cell.font = green_f
                            cell.alignment = center
                        elif col_name in ('URL',):
                            cell.font = Font(size=10)
                            cell.alignment = center
                        elif col_name in ('Lotto', 'Data Asta', 'Città'):
                            cell.font = Font(size=10)
                            cell.alignment = center
                        else:
                            cell.font = Font(size=10)
                            cell.alignment = left_a
                        if r % 2 == 0 and col_name not in ('Prezzo Base', 'Offerta Minima'):
                            cell.fill = alt_fill

                # Link cliccabili
                url_col = col_order.index('URL') + 1
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=url_col)
                    orig = df.iloc[r-2].get('URL', 'N/A')
                    if str(orig).startswith('http'):
                        cell.hyperlink = orig
                        cell.value = "Vedi Asta"
                        cell.font = Font(color="0000CC", underline="single", size=10)
                        cell.alignment = center

                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = f"A1:{chr(ord('A')+len(col_order)-1)}{max_row}"

                # Statistiche
                self._add_stats_sheet(writer, df)

            self._log(f"💾 Salvato: {fpath}", "SUCCESS")
            self._log(f"📊 {len(df)} aste esportate in Excel", "INFO")
            self.root.after(1500, lambda: self._ask_open(fpath))

        except Exception as e:
            import traceback
            self._log(f"❌ Errore salvataggio: {e}\n{traceback.format_exc()}", "ERROR")
            messagebox.showerror("Errore", str(e))

    def _add_stats_sheet(self, writer, df):
        from openpyxl.styles import Font, PatternFill
        rows = [["Statistica", "Valore"],
                ["Totale Aste", len(df)],
                ["Città Diverse", df['Città'].nunique()],
                ["Categorie Diverse", df['Categoria'].nunique()],
                ["", ""], ["TOP CATEGORIE", "N°"]]
        for cat, cnt in df['Categoria'].value_counts().head(10).items():
            rows.append([str(cat), cnt])
        rows += [["", ""], ["TOP CITTÀ", "N°"]]
        for city, cnt in df['Città'].value_counts().head(15).items():
            rows.append([str(city), cnt])
        pd.DataFrame(rows).to_excel(writer, sheet_name='Statistiche', index=False, header=False)
        sw = writer.sheets['Statistiche']
        sw.column_dimensions['A'].width = 35
        sw.column_dimensions['B'].width = 15
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        for cell in sw[1]:
            cell.font = hf
            cell.fill = hfill

    def _ask_open(self, fpath):
        if messagebox.askyesno("File Salvato!",
                               f"✅ Excel salvato!\n📁 {os.path.basename(fpath)}\n\nAprirlo ora?"):
            try:
                if platform.system() == "Windows":
                    os.startfile(fpath)
                elif platform.system() == "Darwin":
                    subprocess.run(["open", fpath])
                else:
                    subprocess.run(["xdg-open", fpath])
            except Exception as e:
                self._log(f"❌ {e}", "ERROR")

    def _update_timer(self):
        if self.timer_running and self.start_time:
            elapsed = time.time() - self.start_time
            h, rem = divmod(int(elapsed), 3600)
            m, s = divmod(rem, 60)
            self.time_lbl.config(text=f"🕐 Tempo: {h:02d}:{m:02d}:{s:02d}")
            if elapsed > 0:
                self.speed_lbl.config(
                    text=f"⚡ Velocità: {(self.total_found/elapsed)*60:.1f} annunci/min")
            self.root.after(1000, self._update_timer)

    def _log(self, msg, level="INFO"):
        icons = {"INFO": "ℹ", "SUCCESS": "✔", "WARNING": "⚠", "ERROR": "✖"}
        ts = datetime.now().strftime("%H:%M:%S")
        formatted = f"[{ts}] {icons.get(level,'ℹ')} {msg}\n"
        def _ins():
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, formatted)
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
        self.root.after(0, _ins)

    def _clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    # ── Mappa interattiva ────────────────────────────────────────────────────
    def _open_mappa(self):
        if not MAP_AVAILABLE:
            messagebox.showwarning(
                "Libreria mancante",
                "La libreria 'tkintermapview' non è installata.\n\n"
                "Aprire il terminale e digitare:\n\n"
                "    pip install tkintermapview\n\n"
                "Poi riavviare l'applicazione.")
            return
        MappaZonaWindow(self.root, self._on_zona_confirmed, self.C,
                        existing_polygon=self.zona_polygon)

    def _on_zona_confirmed(self, comuni, polygon):
        self.zona_comuni = comuni
        self.zona_polygon = polygon
        if comuni:
            nomi = ", ".join(comuni[:5])
            if len(comuni) > 5:
                nomi += f" (+{len(comuni) - 5} altri)"
            self.zona_lbl.config(
                text=f"✔ Zona attiva: {len(comuni)} comuni — {nomi}",
                fg=self.C["success"])
            self._log(f"🗺 Filtro zona attivo: {len(comuni)} comuni selezionati", "SUCCESS")
            for c in comuni:
                self._log(f"   • {c}", "INFO")
        else:
            self.zona_lbl.config(
                text="⚠️ Zona definita ma nessun comune rilevato — verifica la zona",
                fg=self.C["warning"])
            self._log("⚠️ Zona disegnata ma nessun comune riconosciuto", "WARNING")

    def _clear_zona(self):
        self.zona_comuni = []
        self.zona_polygon = []
        self.zona_lbl.config(
            text="Nessun filtro zona attivo — scraping su tutta la provincia",
            fg=self.C["text_dim"])
        self._log("🗺 Filtro zona rimosso", "INFO")

    def _open_folder(self):
        os.makedirs("risultati_aste", exist_ok=True)
        self._open_dir("risultati_aste")

    def _open_debug(self):
        os.makedirs("debug_html_aste", exist_ok=True)
        self._log(f"📁 {os.path.abspath('debug_html_aste')}", "INFO")
        self._open_dir("debug_html_aste")

    def _open_dir(self, d):
        try:
            if platform.system() == "Windows":
                os.startfile(d)
            elif platform.system() == "Darwin":
                subprocess.run(["open", d])
            else:
                subprocess.run(["xdg-open", d])
        except Exception as e:
            messagebox.showerror("Errore", str(e))


# ========== MAIN ==========
def main():
    try:
        if platform.system() == "Windows":
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass
        root = tk.Tk()
        app = AstaLegaleApp(root)

        def on_close():
            if app.is_scraping:
                if messagebox.askokcancel("Esci", "Scraping in corso. Uscire?"):
                    app.scraper.stop_scraping = True
                    root.destroy()
            else:
                root.destroy()

        root.protocol("WM_DELETE_WINDOW", on_close)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("Errore Critico", str(e))


if __name__ == "__main__":
    main()